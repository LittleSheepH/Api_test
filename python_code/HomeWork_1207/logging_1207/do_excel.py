# -*-coding:utf-8-*-
# @Time       :9/12/18上午11:54
# @Autor      :DA BAI CAI
# @Email      :icewong401@163.com
# @File       :do_excel.py
# @Software   :PyCharm
from openpyxl import load_workbook
class do_excel:
    def readdata(self,config_value):
        wb = load_workbook('Api_test_case.xlsx')
        sheet = wb['Sheet1']
        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data={}
            sub_data['case_id']=sheet.cell(i,1).value
            sub_data['url']=sheet.cell(i,2).value
            sub_data['data'] = eval(sheet.cell(i, 3).value)
            sub_data['http_method'] = sheet.cell(i, 4).value
            sub_data['expected'] = str(sheet.cell(i, 5).value)
            test_data.append(sub_data)
        end_data=[]
        if config_value=='all':
            end_data=test_data
        else:
            for item in test_data:
                if item['case_id'] in eval(config_value):
                    end_data.append(item)
        return end_data
if __name__ == '__main__':
    from HomeWork.logging_1207.do_config import Readconfig
    config_value = Readconfig().read_config('do_logging.conf', 'Case', 'button')
    test_data = do_excel().readdata(config_value)
    print(test_data)