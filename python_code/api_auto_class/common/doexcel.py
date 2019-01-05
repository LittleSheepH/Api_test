# -*-coding:utf-8-*-
# @Time       :2018/12/18 23:01
# @Autor      :DA BAI CAI
# @Email      :icewong401@163.com
# @File       :doexcel.py
# @Software   :PyCharm
import openpyxl
import json

import xlrd as xlrd

# from api_auto_class.common.request import Request
from api_auto_class.common.config import Configloader
from api_auto_class.common import concants

class Case:
    def __init__(self):
        self.case_id=None
        self.url=None
        self.data=None
        self.title=None
        self.method =None
        self.excepted = None

class readcase(object):
    def __init__(self,filename):
        try:
            self.file_name=filename
            self.workbook=openpyxl.load_workbook(filename)
        except FileNotFoundError as e:
            raise e
    def get_case(self,sheet_name):
        conf= Configloader()
        sheet=self.workbook[sheet_name]
        max_row=sheet.max_row
        cases=[]
        for r in range(2,max_row+1):
            case=Case()
            case.case_id = sheet.cell(row = r,column = 1).value
            case.url= conf.get('api','url_pre')+sheet.cell(row=r,column=4).value
            case.method=sheet.cell(row=r,column=3).value
            case.data = sheet.cell(row=r, column=5).value
            case.excepted = sheet.cell(row=r, column=6).value
            cases.append(case)
        return cases
    #获取所有的sheetname列表
    def get_sheet_names(self):
        return self.workbook.sheetnames
    def write_actural_by_case_id(self,sheet_name,case_id,actural=None,result=None,amount=None,database=None):
        sheet=self.workbook[sheet_name]
        maxrow= sheet.max_row
        for r in range(2,maxrow+1):#闭区间所以加1
            case_id_r = sheet.cell(row=r,column = 1).value#取第R行的第1列的caseid
            if case_id_r == case_id:
                sheet.cell(r,7).value = actural#写入值
                sheet.cell(r,8).value = result
                sheet.cell(r,9).value = amount
                sheet.cell(r,10).value = database
                self.workbook.save(self.file_name)
                break

    def excel(self, filename):
        if self.file_exist(filename):
            book = xlrd.open_workbook(filename)
            sheet = book.sheet_by_index(0)
            rows = sheet.nrows
            list = []
            for row in range(rows):
                list.append(sheet.row_values(row))
            return list




if __name__ == '__main__':
    read = readcase(concants.data)
    list=read.excel(concants.data)
    print(list)

