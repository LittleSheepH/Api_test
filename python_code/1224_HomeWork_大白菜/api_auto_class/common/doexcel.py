# -*-coding:utf-8-*-
# @Time       :2018/12/18 23:01
# @Autor      :DA BAI CAI
# @Email      :icewong401@163.com
# @File       :doexcel.py
# @Software   :PyCharm
import openpyxl
import json
from api_auto_class.common.request import Request
from api_auto_class.common.config import Configloader

class Case:
    def __init__(self):
        self.case_id=None
        self.url=None
        self.data=None
        self.title=None
        self.method =None
        self.excepted = None

class readcase(object):
    def __init__(self,filename):
        try:
            self.file_name=filename
            self.workbook=openpyxl.load_workbook(filename)
        except FileNotFoundError as e:
            raise e
    def get_case(self,sheet_name):
        conf= Configloader()
        sheet=self.workbook[sheet_name]
        max_row=sheet.max_row
        cases=[]
        for r in range(2,max_row+1):
            case=Case()
            case.case_id = sheet.cell(row = r,column = 1).value
            case.url= conf.get('api','url_pre')+sheet.cell(row=r,column=4).value
            case.method=sheet.cell(row=r,column=3).value
            case.data = sheet.cell(row=r, column=5).value
            case.excepted = sheet.cell(row=r, column=6).value
            cases.append(case)
        return cases
    #获取所有的sheetname列表
    def get_sheet_names(self):
        return self.workbook.sheetnames
    def write_actural_by_case_id(self,sheet_name,case_id,actural=None,result=None,amount=None,database=None):
        sheet=self.workbook[sheet_name]
        maxrow= sheet.max_row
        for r in range(2,maxrow+1):
            case_id_r = sheet.cell(row=r,column = 1).value#取第R行的第1列的caseid
            if case_id_r == case_id:
                sheet.cell(r,7).value = actural#写入值
                sheet.cell(r,8).value = result
                sheet.cell(r,9).value = amount
                sheet.cell(r,10).value = database
                self.workbook.save(self.file_name)
                break


    # def write_result_by_case_id(self,sheetname,case_id,result):
    #     sheet=self.workbook[sheetname]
    #     maxrow= sheet.max_row
    #     # print("***maxrow**{}".format(maxrow))
    #     # print("***case_id**{}".format(case_id))
    #     for r in range(2,maxrow+1):
    #         case_id_r = sheet.cell(row=r,column = 1).value#取第R行的第1列的caseid
    #         # print(sheet.cell(row=r,column = 1).value)
    #         # print("***case_id_r**{}".format(case_id_r))
    #         if case_id_r == case_id:
    #             sheet.cell(r,8).value = result#写入值
    #             self.workbook.save(self.file_name)
    #             break

if __name__ == '__main__':
    do_excel= readcase('../datas/testdatas.xlsx')
    sheetnames=do_excel.get_sheet_names()
    print(sheetnames)
    case_list=['register']
    for sheetname in sheetnames:
        if sheetname in case_list:
            cases =do_excel.get_case(sheetname)#返回的是一个列表
            print('测试用例个数',len(cases))
            for case in cases:#遍历测试用例列表，每个列表元素是一个实例，实例的属性才是要的值，每进for一次，就取一个case实例
                #print("case信息;",case.__dict__)#打印case信息,每个object都有这个属性，打印每个object的值，是一个字典
                data = eval(case.data)
                print(type(data))
                resp = Request(method=case.method,url=case.url,data=data)
                # print(resp.get_statue_code())
                # tesp_test=json.dumps(resp.get_json(),ensure_ascii=False,indent=4)
                # print(tesp_test)
                # print(resp.get_json())
                print(resp.get_statue_code())
                #判断接口响应是否和excel里面的expected的值一致
                print(resp.get_text())
                print(type(case.excepted))



